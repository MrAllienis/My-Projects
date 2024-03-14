import openpyxl
from openpyxl import load_workbook
import pandas as pd
import time


full_all_phrases = []

class Words:

    titles = []
    double_list = []
    words = []
    phrases2 = []
    phrases22 = []
    new_phrases = []
    priority_phrases = []

    final_phrases3 = []
    final_phrases2 = []
    final_phrases1 = []
    final_phrases0 = []


    def game_is_on(self):
        wb = load_workbook(filename='ДобавлениеПредложений.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                self.titles.append(cell)
        wb.close()


    def full_phrases(self):
        wb = load_workbook(filename='ДобавлениеФраз.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                self.final_phrases0.append(cell)
        wb.close()


    def correct_words(self):
        for title in self.titles:
            title = title.replace(',', '')
            title = title.replace(' - ', ' ')
            title = title.lower().split(' ')
            for word in title:
                self.words.append(word)
        # print(self.words)
        self.words = list(set(self.words))
        self.words.sort(key=len, reverse=True)
        # print(self.words)


    def minus(self):
        minus_words = []
        wb = load_workbook(filename='МинусСлова.xlsx')
        sheet = wb['Лист1']
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                minus_words.append(cell)
        wb.close()
        # print(minus_words)
        for word in minus_words:
            if word in self.words:
                self.words.remove(word)
        # self.words.sort(key=len)
        # print(self.words)

    def one_word(self):
        double_word = []
        df = pd.read_excel('Двойные слова.xlsx', header=None,
                           engine='openpyxl')  # header=None, чтобы первая строка не была распознана как заголовок
        data = [list(row) for row in df.itertuples(index=False, name=None)]
        for element in data:
            element = element[0].split(' ')
            self.double_list.append(element)
        # print(self.double_list)

        for el in self.double_list:
            for word in el:
                if word in self.words:
                    double_word.append(word)
                    self.words.remove(word)
            double_word = ("/".join(double_word))
            # print(double_word)
            self.words.append(double_word)
            double_word = []
        # print(self.words)


    def generate_phrases(self):
        for word1 in self.words:
            for word2 in self.words:
                for word3 in self.words:
                    if word1[:-1] != word2[:-1] and word1[:-1] != word3[:-1] and word2[:-1] != word3[:-1]:
                        self.phrases2.append(f"{word1} {word2} {word3}")
        # print(self.phrases2)
        # print(len(self.phrases2))


    def generate_phrases2(self):
        for word1 in self.words:
            for word2 in self.words:
                if word1[:-1] != word2[:-1]:
                    self.phrases22.append(f"{word1} {word2}")
        # print(self.phrases22)
        # print(len(self.phrases22))


    def validation(self):
        for phrase1 in self.phrases2:
            for phrase2 in self.phrases2:
                if (phrase1.split(' ')[0] == phrase2.split(' ')[1]) and (phrase1.split(' ')[1] == phrase2.split(' ')[0]):
                    # print(f"{phrase1} {phrase2}")
                    self.phrases2.remove(phrase2)
                if (phrase1.split(' ')[1] == phrase2.split(' ')[2]) and (phrase1.split(' ')[2] == phrase2.split(' ')[1]):
                    # print(f"{phrase1} {phrase2}")
                    self.phrases2.remove(phrase2)
                if (phrase1.split(' ')[0] == phrase2.split(' ')[2]) and (phrase1.split(' ')[2] == phrase2.split(' ')[0]):
                    # print(f"{phrase1} {phrase2}")
                    self.phrases2.remove(phrase2)
                if (phrase1.split(' ')[1] == phrase2.split(' ')[0]) and (phrase1.split(' ')[2] == phrase2.split(' ')[1]):
                #     # print(f"{phrase1} {phrase2}")
                    self.phrases2.remove(phrase2)


    def validation2(self):
        for phrase1 in self.phrases22:
            for phrase2 in self.phrases22:
                if (phrase1.split(' ')[0] == phrase2.split(' ')[1]) and (phrase1.split(' ')[1] == phrase2.split(' ')[0]):
                    self.phrases22.remove(phrase2)
        # print(self.phrases22)
        # print(len(self.phrases22))



        # for phrase in self.phrases2:
        #     self.new_phrases.append(phrase)
        # print(self.new_phrases)
        # print(len(self.new_phrases))
        self.phrases2.sort(key=len, reverse=True)
        # print(self.phrases2)
        # print(len(self.phrases2))


    def priority(self):
        priority_words = []
        wb = load_workbook(filename='ПриоритетныеСлова.xlsx')
        sheet = wb['Лист1']
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                priority_words.append(cell)
        wb.close()
        self.phrases2.sort(key=len, reverse=True)
        for phrase in self.phrases2:
            if (phrase.split(' ')[0] in priority_words) or (phrase.split(' ')[1] in priority_words) or (phrase.split(' ')[2] in priority_words):
                self.priority_phrases.append(phrase)
                self.priority_phrases.sort()
        self.phrases2 = self.priority_phrases
        self.phrases2.sort(key=len, reverse=False)
        # print(self.phrases2)
        # print(len(self.phrases2))

    def priority2(self):
        priority_words = []
        wb = load_workbook(filename='ПриоритетныеСлова.xlsx')
        sheet = wb['Лист1']
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                priority_words.append(cell)
        wb.close()
        self.phrases22.sort(key=len, reverse=True)
        for phrase in self.phrases22:
            if (phrase.split(' ')[0] in priority_words) or (phrase.split(' ')[1] in priority_words):
                self.priority_phrases.append(phrase)
                self.priority_phrases.sort()
        self.phrases22 = self.priority_phrases
        self.phrases22.sort(key=len, reverse=False)
        # print(self.phrases22)
        # print(len(self.phrases22))


    def priority1(self):
        priority_words = []
        wb = load_workbook(filename='ПриоритетныеСлова.xlsx')
        sheet = wb['Лист1']
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                priority_words.append(cell)
        wb.close()
        for phrase in self.words:
            if phrase in priority_words:
                self.priority_phrases.append(phrase)
                self.priority_phrases.sort()
        self.words = self.priority_phrases
        self.words.sort(key=len, reverse=False)


    def final3(self):
        for phrase in self.phrases2:
            phrase = phrase.replace('/', ' ')
            self.final_phrases3.append(phrase)
        print('')
        print(f'Список фраз из 3-ёх слов. Количество фраз - {len(self.final_phrases3)}')
        print(self.final_phrases3)



    def final2(self):
        for phrase in self.phrases22:
            phrase = phrase.replace('/', ' ')
            self.final_phrases2.append(phrase)
        print('')
        print(f'Список фраз из 2-ух слов. Количество фраз - {len(self.final_phrases2)}')
        print(self.final_phrases2)




    def dont_find(self):
        print('')
        input("Пожалуйста, отредактируйте файл 'НеИскатьЧастотность.xlsx', после чего сохраните его, а затем нажмите клавишу 'Enter'.")
        minus_phrases = []
        wb = load_workbook(filename='НеИскатьЧастотность.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                minus_phrases.append(cell)
        wb.close()
        # print(minus_words)
        for phrase in minus_phrases:
            if phrase in full_all_phrases:
                full_all_phrases.remove(phrase)
        # self.words.sort(key=len)
        # print(self.words)


    def null(self):
        self.double_list = []
        self.words = []
        self.phrases2 = []
        self.phrases22 = []
        self.new_phrases = []
        self.priority_phrases = []
        self.correct_words()
        self.minus()
        self.one_word()


    def main(self):
        print("Вас приветствует парсер WordStat!")
        variant = None
        while variant not in ('1', '2'):
            variant = str(input("Нажмите цифру 1 для генерации фраз из предложений, или цифру 2 для поиска по целым фразам: "))
            if variant == '1':
                input("Перед началом работы убедитесь, что в файл 'ДобавлениеПредложений.xlsx' добавлены необходимые предложения. Затем нажмите клавишу 'Enter'. ")
            elif variant == '2':
                input("Перед началом работы убедитесь, что в файл 'ДобавлениеФраз.xlsx' добавлены необходимые фразы. Затем нажмите клавишу 'Enter'. ")
            else:
                print(f"Вы ввели недопустимое значение: {variant}")

        if variant == '1':
            self.game_is_on()

            max_count_words_in_phrases = None
            with_priority = None
            while max_count_words_in_phrases not in ('1', '2', '3'):
                max_count_words_in_phrases = str(input("Введите максимальное количество слов во фразе для генерации (1 / 2 / 3): "))
                if max_count_words_in_phrases not in ('1', '2', '3'):
                    print(f"Вы ввели недопустимое значение: {max_count_words_in_phrases}")

            while (with_priority != True) and (with_priority != False):
                with_priority = str(input("Нужно ли учитывать приоритетные слова? Введите цифру 1 - если да, цифру 0 - если нет: "))
                if with_priority == '1':
                    with_priority = True
                elif with_priority == '0':
                    with_priority = False
                else:
                    print(f"Вы ввели недопустимое значение: {with_priority}")

            self.correct_words()
            self.minus()
            self.one_word()

            if max_count_words_in_phrases == '3':
                self.generate_phrases()
                self.validation()
                if with_priority == True:
                    self.priority()

                self.final3()
                self.null()

                self.generate_phrases2()
                self.validation2()
                if with_priority == True:
                    self.priority2()

                self.final2()
                self.null()

                if with_priority == True:
                    self.priority1()
                for phrase in self.words:
                    phrase = phrase.replace('/', ' ')
                    self.final_phrases1.append(phrase)
                print('')
                print(f'Список фраз из 1-го слова. Количество фраз - {len(self.final_phrases1)}')
                print(self.final_phrases1)

            elif max_count_words_in_phrases == '2':
                self.generate_phrases2()
                self.validation2()
                if with_priority == True:
                    self.priority2()

                self.final2()
                self.null()


                if with_priority == True:
                    self.priority1()
                for phrase in self.words:
                    phrase = phrase.replace('/', ' ')
                    self.final_phrases1.append(phrase)
                print('')
                print(f'Список фраз из 1-го слова. Количество фраз - {len(self.final_phrases1)}')
                print(self.final_phrases1)

            elif max_count_words_in_phrases == '1':
                if with_priority == True:
                    self.priority1()
                for phrase in self.words:
                    phrase = phrase.replace('/', ' ')
                    self.final_phrases1.append(phrase)
                print('')
                print(f'Список фраз из 1-го слова. Количество фраз - {len(self.final_phrases1)}')
                print(self.final_phrases1)

            full_all_phrases.extend(self.final_phrases1)
            full_all_phrases.extend(self.final_phrases2)
            full_all_phrases.extend(self.final_phrases3)

            # print(full_all_phrases)
            self.dont_find()
            print('')
            print('Финальный список фраз для отправления в WordStat: ')
            print(full_all_phrases)

        elif variant == '2':
            self.full_phrases()
            full_all_phrases.extend(self.final_phrases0)
            print('')
            print('Финальный список фраз для отправления в WordStat: ')
            print(full_all_phrases)


user1 = Words()
user1.main()
