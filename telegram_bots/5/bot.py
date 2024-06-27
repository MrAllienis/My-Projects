import telebot
from telebot import types
from parser1 import tickets
import time
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font


bot = telebot.TeleBot('TOKEN')
stop = False
count_of_tickets = 0

@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, f'Приветствуем Вас, {message.from_user.full_name}.\nМеню бота: ', reply_markup=get_keyboard('main'))


def get_keyboard(key):
    keyboard = types.InlineKeyboardMarkup()
    if key == 'main':
        keyboard.add(types.InlineKeyboardButton('Начать парсинг билетов', callback_data='parsing'))
        keyboard.add(types.InlineKeyboardButton('Очистить базу данных', callback_data='clear'))
        keyboard.add(types.InlineKeyboardButton('Показать статистику', callback_data='statistic'))
    elif key == 'stop':
        keyboard.add(types.InlineKeyboardButton('STOP', callback_data='stop'))
    elif key == 'clear':
        keyboard.add(types.InlineKeyboardButton('ДА', callback_data='yes'))
        keyboard.add(types.InlineKeyboardButton('НЕТ', callback_data='no'))
    return keyboard


def fast_statistic():
    statistic, fast_stat = tickets.get_statistic()
    message = f'Быстрая статистика из 10-и самых частых цифр:\n\n'
    for num in fast_stat:
        message += f'Число {num[0]}:   <b>{num[1]}</b> совпадений\n'
    message += '\nПолная статистика содержится в файле ниже:'
    return message, statistic


@bot.callback_query_handler(func=lambda call: True)
def common_button(call):
    global stop, count_of_tickets
    stop = False
    if call.data == 'parsing':
        bot.send_message(call.from_user.id, f'Происходит парсинг, для остановки - нажмите кнопку ниже: ', reply_markup=get_keyboard('stop'))
        while stop is False:
            try:
                tickets.parse_tickets()
                count_of_tickets += 20
            except:
                bot.send_message(call.from_user.id, 'Ошибка сервера, запрос не отправился, пробуем еще раз')
            time.sleep(1)
    elif call.data == 'stop':
        stop = True
        bot.send_message(call.from_user.id, f'Парсинг завершён ({count_of_tickets} билетов за эту сессию. {tickets.count()} за всё время).  Вы вернулись в главное меню', reply_markup=get_keyboard('main'))
        count_of_tickets = 0
    elif call.data == 'clear':
        bot.send_message(call.from_user.id, f'Вы уверены, что хотите очистить все данные?', reply_markup=get_keyboard('clear'))
    elif call.data =='yes':
        tickets.clear_all_data()
        bot.send_message(call.from_user.id, f'База данных полностью очищена. Вы вернулись в главное меню', reply_markup=get_keyboard('main'))
    elif call.data =='no':
        bot.send_message(call.from_user.id, f'Вы вернулись в главное меню', reply_markup=get_keyboard('main'))
    elif call.data =='statistic':
        text, statistic = fast_statistic()
        wb = Workbook()
        ws = wb.active
        ws.append(["Число", "Количество"])
        ws.column_dimensions['B'].width = 12
        for cell in ws["1:1"]:
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(border_style="thin"))
        for row in statistic:
            ws.append(row)
        wb.save("Полная статистика.xlsx")
        bot.send_message(call.from_user.id, text=text, parse_mode='html')
        with open("Полная статистика.xlsx", 'rb') as file:
            bot.send_document(call.from_user.id, file)
        bot.send_message(call.from_user.id, 'Вы вернулись в главное меню', reply_markup=get_keyboard('main'))


while True:
    try:
        bot.polling(non_stop=True, interval=0)
    except:
        time.sleep(2)


