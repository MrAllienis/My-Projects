import requests
import random
import pandas as pd
import os
import re
import subprocess
import psycopg2
import time
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from download_file import get_csv



DB_NAME = "postgres1"
DB_USER = "postgres"
DB_PASSWORD = "1a2s3d4f"
DB_HOST = "localhost"
DB_PORT = "5432"

LAST_SHOP_PRODUCT = ''
LAST_SHOP_COUPON = ''

def get_campaign(campaign_id):
    access_token = 'dHvn1JxM8B6sg113IGjAKHpbc3IJlj'

    # Параметры запроса
    params = {
        # 'region': 'RU',  # Попробуйте изменить на другой регион, например, 'RU'
        'limit': 500,  # Увеличьте лимит для получения большего количества купонов
        # 'language': 'ru',
    }

    # Заголовки запроса
    headers = {
        'Authorization': f'Bearer {access_token}'
    }
    # Выполнение GET-запроса
    url = f'https://api.admitad.com/advcampaigns/{campaign_id}/website/2684875/'
    correct = ''
    while correct != 200:
        response = requests.get(url, headers=headers, params=params)
        print(response.status_code)
        correct = response.status_code
        if response.status_code == 200:
            data = response.json()
            # print(data)
            ref_link = data['gotolink']
            # ic(data['products_csv_link'])
            # ic(data['products_xml_link'])
            legal_info = data['advertiser_legal_info']
            # ic(data['description'])
            return ref_link, legal_info
        else:
            print('Ошибка при выполнении запроса:', response.status_code, response.text)
            time.sleep(5)
        # except:
        #     time.sleep(5)
        #     print('Error')


    # Проверка ответа


def get_image_xlsx(campaign_name):
    # Загружаем файл Excel
    file_path = 'Shop-Photo.xlsx'
    wb = load_workbook(filename=file_path)
    # Выбираем активный лист (или можно указать конкретное имя листа)
    ws = wb.active

    # Перебираем строки в листе
    for row in ws.iter_rows(values_only=True):
        # Проверяем значение в первой колонке
        if row[0] == campaign_name:
            filtered_row = [cell for cell in row if cell is not None]
            # Возвращаем строку целиком, если значение совпадает
            try:
                image = random.choice(filtered_row[1:])
                return image
            except:
                return None
    # Если значение не найдено, возвращаем None
    return None


def select_products(post_link=False):
    global LAST_SHOP_PRODUCT
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    if post_link is False:
        # SQL запрос для получения случайной строки
        # def get_quary():
        #     cursor.execute("""SELECT sales_app_shop.shop_id
        #                 FROM sales_app_shop
        #                 WHERE (legal_info, goto_link) IS NOT NULL
        #                 ORDER BY RANDOM()
        #                 LIMIT 1
        #                 """)
        #     shop_id = cursor.fetchone()[0]
        #
        #     cursor.execute("""SELECT sales_app_product.category1
        #             FROM sales_app_product
        #             INNER JOIN sales_app_shop
        #             ON sales_app_product.shop_id = sales_app_shop.shop_id
        #             WHERE (sales_app_product.category1, sales_app_product.image, sales_app_shop.name, sales_app_product.name, price, url, legal_info, goto_link) IS NOT NULL
        #             AND sales_app_product.used IS FALSE
        #             AND sales_app_product.shop_id = %s
        #             ORDER BY RANDOM()
        #             LIMIT 1
        #             """, (shop_id,))
        #     return cursor.fetchone()
        #
        # search_name = get_quary()
        # while search_name is None:
        #     search_name = get_quary()
        # search_name = search_name[0]
        #
        # query = """
        #         SELECT sales_app_product.image, sales_app_shop.name, sales_app_product.name, price, old_price, url, legal_info, goto_link, id
        #         FROM sales_app_product
        #         INNER JOIN sales_app_shop
        #         ON sales_app_product.shop_id = sales_app_shop.shop_id
        #         WHERE
        #             sales_app_product.price = (
        #                 SELECT MIN(price)
        #                 FROM sales_app_product
        #                 WHERE category1 = %s
        #             )
        #             AND sales_app_product.category1 = %s
        #             AND (sales_app_product.image, sales_app_shop.name, sales_app_product.name, price, url, legal_info, goto_link, category1) IS NOT NULL
        #             AND sales_app_product.used IS FALSE
        #         ORDER BY RANDOM()
        #         LIMIT 1;
        #         """
        # # Выполнение запроса
        # cursor.execute(query, (search_name, search_name))
        # # Получение результата
        # result = cursor.fetchone()
        for_search = ['Женская одежда', 'Женская обувь', 'Мужская одежда', 'Мужская обувь', 'Детская одежда',
                      'Детская обувь', 'Телефоны', 'Смартфоны', 'Смарт-часы', 'Ноутбуки', 'Планшеты', 'Телевизоры',
                      'Умный дом', 'Беспроводные наушники', 'Бытовая техника', 'Электроника']
        category_element = random.choice(for_search)
        print(category_element)
        cursor.execute("""
                        SELECT sales_app_product.image, sales_app_shop.name, sales_app_product.name, price, old_price, url, legal_info, goto_link, id
                        FROM sales_app_product
                        INNER JOIN sales_app_shop
                        ON sales_app_product.shop_id = sales_app_shop.shop_id
                        WHERE 
                            sales_app_product.category1 ILIKE %s
                            AND (sales_app_product.image, sales_app_shop.name, sales_app_product.name, price, url, legal_info, goto_link, category1) IS NOT NULL
                            AND sales_app_product.used IS FALSE
                            AND sales_app_shop.name <> %s
                            AND sale >= 50
                        ORDER BY RANDOM()
                        LIMIT 1;
                        """, ('%' + category_element + '%', LAST_SHOP_PRODUCT))
        result = cursor.fetchone()
        LAST_SHOP_PRODUCT = result[1]
    else:
        cursor.execute("""SELECT sales_app_product.image, sales_app_shop.name, sales_app_product.name, price, old_price, url, legal_info, goto_link, id
                FROM sales_app_product
                INNER JOIN sales_app_shop
                ON sales_app_product.shop_id = sales_app_shop.shop_id
                WHERE sales_app_product.url = %s""", (post_link,))
        result = cursor.fetchone()

    # Закрытие курсора и подключения
    cursor.close()
    connection.close()

    return result[0], result[1], result[2], result[3], result[4], result[5], result[6], result[7], result[8]


def select_from_db(promo=False, name_promo=False):
    global LAST_SHOP_COUPON
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    if name_promo is False:
        if promo is False:
            # SQL запрос для получения случайной строки
            query = """
                    SELECT *
                    FROM sales_app_coupon
                    INNER JOIN sales_app_shop
                    ON sales_app_coupon.shop_id = sales_app_shop.shop_id
                    WHERE discount IS NOT NULL
                    AND sales_app_coupon.name IS NOT NULL
                    AND legal_info IS NOT NULL
                    AND sales_app_coupon.used = FALSE
                    AND sales_app_shop.shop_id <> '20158'
                    AND sales_app_coupon.code IN ('NOT REQUIRED', 'NOT REQUIRE', 'НЕ НУЖЕН', 'Не нужен')
                    AND discount LIKE '%\%%'  
                    ORDER BY 
                        CASE 
                            WHEN CAST(REPLACE(discount, '%', '') AS INTEGER) >= 30 THEN 0
                            ELSE 1
                        END,
                        RANDOM()
                    LIMIT 1;
                """
        else:
            # SQL запрос для получения случайной строки
            query = """
                    SELECT *
                    FROM sales_app_coupon
                    INNER JOIN sales_app_shop
                    ON sales_app_coupon.shop_id = sales_app_shop.shop_id
                    WHERE discount IS NOT NULL
                    AND sales_app_coupon.name IS NOT NULL
                    AND legal_info IS NOT NULL
                    AND sales_app_coupon.used = FALSE
                    AND sales_app_shop.shop_id <> '20158'
                    AND sales_app_coupon.code NOT IN ('NOT REQUIRED', 'NOT REQUIRE', 'НЕ НУЖЕН', 'Не нужен')
                    AND discount LIKE '%\%%'  
                    ORDER BY 
                        CASE 
                            WHEN CAST(REPLACE(discount, '%', '') AS INTEGER) >= 30 THEN 0
                            ELSE 1
                        END,
                        RANDOM()
                    LIMIT 1;
                    """
        # Выполнение запроса
        try:
            cursor.execute(query)
            # Получение результата
            result = cursor.fetchone()

            LAST_SHOP_COUPON = result[13]

            if result[14] == 25179:
                connection.commit()
                cursor.close()
                connection.close()
                cursor.execute("""UPDATE sales_app_coupon SET used = TRUE WHERE coupon_id = %s""", (result[0],))
                return result[0], result[13], result[2], result[3], '', result[5], result[6], result[12], result[17], result[7]
            else:
                cursor.execute("""UPDATE sales_app_coupon SET used = TRUE WHERE coupon_id = %s""", (result[0],))
                connection.commit()
                # Закрытие курсора и подключения
                cursor.close()
                connection.close()
        except Exception as e:
            cursor.execute("""UPDATE sales_app_coupon SET used = TRUE WHERE coupon_id = %s""", (result[0],))
            # Закрытие курсора и подключения
            cursor.close()
            connection.close()

    else:
        # Выполнение запроса
        try:
            cursor.execute("""
                    SELECT *
                    FROM sales_app_coupon
                    INNER JOIN sales_app_shop
                    ON sales_app_coupon.shop_id = sales_app_shop.shop_id
                    WHERE sales_app_coupon.link = %s;
                    """, (name_promo,))
            # Получение результата
            result = cursor.fetchone()

        finally:
            connection.commit()
            # Закрытие курсора и подключения
            cursor.close()
            connection.close()


    return result[0], result[13], result[2], result[3], result[4], result[5], result[6], result[12], result[17], result[7]



def update_used_product(id, chat_id, message_id):
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    # Выполнение запроса
    try:
        cursor.execute("""
                        UPDATE sales_app_product
                        SET used = TRUE, chat_id = %s, message_id = %s
                        WHERE id = %s;
                        """, (chat_id, message_id, id,))
        connection.commit()
    finally:
        # Закрытие курсора и подключения
        cursor.close()
        connection.close()



def update_used_coupon(coupon_id, chat_id, message_id):
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    # Выполнение запроса
    try:
        cursor.execute("""
                        UPDATE sales_app_coupon
                        SET used = TRUE, chat_id = %s, message_id = %s
                        WHERE coupon_id = %s;
                        """, (chat_id, message_id, coupon_id,))
        connection.commit()
    finally:
        # Закрытие курсора и подключения
        cursor.close()
        connection.close()


def delete_old_coupons():
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    # Выполнение запроса
    try:
        cursor.execute("""DELETE FROM sales_app_coupon WHERE end_start < %s AND used = FALSE; """, (datetime.now(),))


        cursor.execute(""" DELETE FROM sales_app_coupon WHERE end_start < %s AND used = TRUE AND (chat_id, message_id) IS NOT NULL
                        RETURNING chat_id, message_id; """,
                       (datetime.now(),))
        messages_for_deleting = cursor.fetchall()
        connection.commit()
    finally:
        # Закрытие курсора и подключения
        cursor.close()
        connection.close()
    try:
        return messages_for_deleting
    finally:
        print(f"Старые записи успешно удалены из таблицы.")

def save_store_to_xlsx(shops_names):
    file_path = 'Shop-Photo.xlsx'
    # Открываем файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Читаем существующие значения из первой колонки
    existing_values = set()
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        if row[0] is not None:
            existing_values.add(row[0])

    # Добавляем новые элементы, если их нет в первой колонке
    for item in shops_names:
        if item not in existing_values:
            sheet.append([item])
            existing_values.add(item)

    # Сохраняем изменения
    workbook.save(file_path)



def save_shops_only(data):
    # Подключение к базе данных
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    # SQL-запрос для вставки данных
    shops_names = set()
    categories_shop(data)
    for shop in data['results']:
        if shop['gotolink'] != '':
            shops_names.add(shop['name'])
            # Выполнение запроса
            try:
                cursor.execute(
                    """
                    INSERT INTO sales_app_shop (shop_id, name, goto_link, legal_info)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (shop_id)
                    DO UPDATE SET
                        goto_link = EXCLUDED.goto_link,
                        legal_info = EXCLUDED.legal_info
                    """,
                    (shop['id'], shop['name'], shop['gotolink'], shop['advertiser_legal_info'])
                )
            except Exception as e:
                print(f'Магазин {shop["name"]} не сохранен - {e}')
            for cat in shop['categories']:
                if cat['name'] not in ['11.11', 'SME']:
                    try:
                        cursor.execute("""INSERT INTO sales_app_shop_category (shop_id, categoryshop_id) VALUES (%s, %s) ON CONFLICT (shop_id, categoryshop_id) DO NOTHING""", (shop['id'], cat['id']))
                    except Exception as e:
                        print(f"Связь {shop['name']} + {cat['name']} не сохранена. {e}")
    # Фиксация изменений
    connection.commit()
    cursor.close()
    connection.close()

    save_store_to_xlsx(shops_names)


def save_products_to_db(data):
    # Подключение к базе данных
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    products_for_delete = []
    # SQL-запрос для вставки данных
    shops_names = set()
    not_rub_shops = []
    categories_shop(data)
    for shop in data['results']:
        if (shop['gotolink'] not in ['', None]) and (shop['currency'] == 'RUB' and shop['id'] != 25179 and shop['name'] == 'rendez-vous.ru'):
            shops_names.add(shop['name'])
            if (shop['categories'][0]['name'] != 'Интернет-магазины') or (len(shop['categories']) == 1):
                category = shop['categories'][0]
            else:
                category = shop['categories'][1]
            # Выполнение запроса
            try:
                cursor.execute(
                    """SELECT last_update FROM sales_app_shop WHERE shop_id = %s;""", (shop['id'],))
                last_update = cursor.fetchone()[0]
                if last_update is None:
                    last_update = datetime.fromisoformat(shop['feeds_info'][0]['admitad_last_update']).replace(
                        year=2000)
                try:
                    cursor.execute(
                        """INSERT INTO sales_app_shop (shop_id, name, goto_link, legal_info) 
                                    VALUES (%s, %s, %s, %s) ON CONFLICT (shop_id) DO NOTHING """,
                        (shop['id'], shop['name'], shop['gotolink'],
                         shop['advertiser_legal_info']))
                except Exception as e:
                    print(f'Магазин {shop["name"]} не сохранен - {e}')
                if datetime.fromisoformat(shop['feeds_info'][0]['admitad_last_update']) > last_update:
                    if shop['products_csv_link'] not in ['', None]:
                        try:
                            cursor.execute("SELECT id, price FROM sales_app_product WHERE shop_id = %s;", (shop['id'],))
                            current_rows = {row[0].split('_%f%_')[0]: row[1] for row in cursor.fetchall()}
                            current_ids = set(current_rows.keys())
                            print(f"Скачиваются товары магазина {shop['name']}. Это может занять несколько минут")
                            file_path = 'data.csv'
                            # if os.path.exists(file_path):
                            #     os.remove(file_path)
                            if shop['id'] not in [25179]:
                                products = get_csv(shop['feeds_info'])
                                new_rows = {str(row[0]): row[1] for row in list(products.values)}
                                new_ids = set(str(row[0]) for row in list(products.values))
                                # Определение данных для удаления
                                ids_to_delete = current_ids - new_ids
                                for key in current_rows:
                                    # Проверяем, если ключ существует во втором словаре и значения не совпадают
                                    if (key in new_rows):
                                        new_rows[key] = re.sub(r'[^0-9.]', '', str(new_rows[key]))
                                        try:
                                            new_rows[key] = new_rows[key].replace(',', '.')
                                        except:
                                            pass
                                        try:
                                            if float(current_rows[key]) != float(new_rows[key]):
                                                ids_to_delete.add(key)
                                        except:
                                            ids_to_delete.add(key)
                                # Удаление данных, которых нет в новых данных или цена которых не совпадает
                                if ids_to_delete:
                                    ids_to_delete = set(f'{row}_%f%_{shop["id"]}' for row in ids_to_delete)
                                    cursor.execute(
                                        """DELETE FROM sales_app_product WHERE id IN %s RETURNING chat_id, message_id; """,
                                        (tuple(ids_to_delete),))
                                    for fact_used in cursor.fetchall():
                                        if (fact_used[0] and fact_used[1]) is not None:
                                            products_for_delete.append(fact_used)
                                current_rows.clear()
                                current_ids.clear()
                                new_rows.clear()
                                new_ids.clear()
                                products = products[0:0]
                                ids_to_delete.clear()
                                file_path = 'data.csv'
                                delimiter = ';'
                                print('Добавление товаров в таблицу')
                                try:
                                    selected_columns = ['id', 'name', 'categoryId', 'price', 'oldprice', 'picture',
                                                        'url', 'currencyId']  # Укажите нужные колонки
                                    for result in pd.read_csv(file_path, sep=delimiter, chunksize=10000,
                                                              usecols=selected_columns, on_bad_lines='skip'):
                                        result = result[selected_columns]
                                        for product in result.values:
                                            if product[7].upper() in ['RUB', 'RUR', 'RU']:
                                                try:
                                                    try:
                                                        cat1 = str(product[2])
                                                    except:
                                                        cat1 = ''
                                                    try:
                                                        product[3] = product[3].replace(',', '.')
                                                    except:
                                                        pass
                                                    product[3] = re.sub(r'[^0-9.]', '', str(product[3]))
                                                    if (str(product[3]) and product[6]) not in ['NaN', 'nan', None]:
                                                        if float(product[3]) > 0:
                                                            try:
                                                                product[4] = product[4].replace(',', '.')
                                                            except:
                                                                pass
                                                            if product[4] is not None:
                                                                product[4] = re.sub(r'[^0-9.]', '', str(product[4]))

                                                            if product[4] == '':
                                                                product[4] = 0
                                                            try:
                                                                if str(product[4]) in ['NaN', 'nan']:
                                                                    product[4] = 0
                                                            except:
                                                                product[4] = 0
                                                            try:
                                                                product[4] = float(product[4])
                                                            except:
                                                                product[4] = 0
                                                            try:
                                                                try:
                                                                    if (product[4] not in ['NaN', 'nan', None, 0]) and (
                                                                            product[3] not in ['NaN', 'nan', None,
                                                                                               0]) and (
                                                                            float(product[4]) > float(product[3])):
                                                                        sale = round((float(product[4]) - float(
                                                                            product[3])) / float(product[4]) * 100)
                                                                    else:
                                                                        sale = 0
                                                                except:
                                                                    sale = 0
                                                                super_id = f'{str(product[0])}_%f%_{str(shop["id"])}'
                                                                cursor.execute(
                                                                    """INSERT INTO sales_app_product (id, name, price, old_price, sale, image, url, category1, shop_id, used) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE) ON CONFLICT (id) DO NOTHING """,
                                                                    (super_id, product[1], product[3], product[4], sale,
                                                                     product[5],
                                                                     product[6], cat1, shop['id'],))
                                                                # Фиксация транзакции
                                                                # connection.commit()
                                                            except Exception as e:
                                                                print(f'Ошибка {e} товара {product}')
                                                                if connection:
                                                                    connection.rollback()
                                                except Exception as e:
                                                    print(e)
                                except:
                                    selected_columns = ['id', 'name', 'categoryId', 'price', 'picture', 'url',
                                                        'currencyId']  # Укажите нужные колонки
                                    try:
                                        for result in pd.read_csv(file_path, sep=delimiter, chunksize=10000,
                                                                  usecols=selected_columns, on_bad_lines='skip'):
                                            result['oldprice'] = 0
                                            correct_columns = ['id', 'name', 'categoryId', 'price', 'oldprice',
                                                               'picture', 'url', 'currencyId']
                                            result = result[correct_columns]
                                            for product in result.values:
                                                if product[7].upper() in ['RUB', 'RUR', 'RU']:
                                                    try:
                                                        try:
                                                            cat1 = str(product[2])
                                                        except:
                                                            cat1 = ''
                                                        try:
                                                            product[3] = product[3].replace(',', '.')
                                                        except:
                                                            pass
                                                        product[3] = re.sub(r'[^0-9.]', '', str(product[3]))
                                                        if (str(product[3]) and product[6]) not in ['NaN', 'nan', None]:
                                                            if float(product[3]) > 0:
                                                                try:
                                                                    sale = 0
                                                                    super_id = f'{str(product[0])}_%f%_{str(shop["id"])}'
                                                                    cursor.execute(
                                                                        """INSERT INTO sales_app_product (id, name, price, old_price, sale, image, url, category1, shop_id, used) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE) ON CONFLICT (id) DO NOTHING """,
                                                                        (super_id, product[1], product[3], product[4],
                                                                         sale, product[5], product[6], cat1,
                                                                         shop['id'],))
                                                                    # Фиксация транзакции
                                                                    # connection.commit()
                                                                except Exception as e:
                                                                    print(f'Ошибка {e} товара {product}')
                                                                    if connection:
                                                                        connection.rollback()
                                                    except Exception as e:
                                                        print(e)
                                    except:
                                        selected_columns = ['id', 'name', 'price', 'picture', 'url', 'currencyId']
                                        try:
                                            for result in pd.read_csv(file_path, sep=delimiter, chunksize=10000,
                                                                      usecols=selected_columns, on_bad_lines='skip'):
                                                result['categoryId'] = 'товар'
                                                result['oldprice'] = 0
                                                correct_columns = ['id', 'name', 'categoryId', 'price', 'oldprice',
                                                                   'picture', 'url', 'currencyId']
                                                result = result[correct_columns]
                                                for product in result.values:
                                                    if product[7].upper() == 'RUB':
                                                        try:
                                                            try:
                                                                cat1 = str(product[2])
                                                            except:
                                                                cat1 = ''
                                                            try:
                                                                product[3] = product[3].replace(',', '.')
                                                            except:
                                                                pass
                                                            product[3] = re.sub(r'[^0-9.]', '', str(product[3]))
                                                            if (str(product[3]) and product[6]) not in ['NaN', 'nan',
                                                                                                        None]:
                                                                if float(product[3]) > 0:
                                                                    try:
                                                                        sale = 0
                                                                        super_id = f'{str(product[0])}_%f%_{str(shop["id"])}'
                                                                        cursor.execute(
                                                                            """INSERT INTO sales_app_product (id, name, price, old_price, sale, image, url, category1, shop_id, used) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE) ON CONFLICT (id) DO NOTHING """,
                                                                            (super_id, product[1], product[3],
                                                                             product[4], sale, product[5], product[6],
                                                                             cat1, shop['id'],))
                                                                        # Фиксация транзакции
                                                                        # connection.commit()
                                                                    except Exception as e:
                                                                        print(f'Ошибка {e} товара {product}')
                                                                        if connection:
                                                                            connection.rollback()
                                                        except Exception as e:
                                                            print(e)
                                        except:
                                            with open(file_path, 'rb') as rawdata:
                                                result = chardet.detect(
                                                    rawdata.read(100000))  # Читаем часть файла для анализа
                                                encoding = result['encoding']
                                                print(f"Определенная кодировка файла: {encoding}")
                                            for result in pd.read_csv(file_path, sep=delimiter, chunksize=10000,
                                                                      usecols=selected_columns, encoding=encoding,
                                                                      on_bad_lines='skip'):
                                                result['categoryId'] = 'товар'
                                                result['oldprice'] = 0
                                                correct_columns = ['id', 'name', 'categoryId', 'price', 'oldprice',
                                                                   'picture', 'url', 'currencyId']
                                                result = result[correct_columns]
                                                for product in result.values:
                                                    if product[7].upper() == 'RUB':
                                                        try:
                                                            try:
                                                                cat1 = str(product[2])
                                                            except:
                                                                cat1 = ''
                                                            try:
                                                                product[3] = product[3].replace(',', '.')
                                                            except:
                                                                pass
                                                            product[3] = re.sub(r'[^0-9.]', '', str(product[3]))
                                                            if (str(product[3]) and product[6]) not in ['NaN', 'nan',
                                                                                                        None]:
                                                                if float(product[3]) > 0:
                                                                    try:
                                                                        sale = 0
                                                                        super_id = f'{str(product[0])}_%f%_{str(shop["id"])}'
                                                                        cursor.execute(
                                                                            """INSERT INTO sales_app_product (id, name, price, old_price, sale, image, url, category1, shop_id, used) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE) ON CONFLICT (id) DO NOTHING """,
                                                                            (super_id, product[1], product[3],
                                                                             product[4], sale, product[5], product[6],
                                                                             cat1, shop['id'],))
                                                                        # Фиксация транзакции
                                                                        # connection.commit()
                                                                    except Exception as e:
                                                                        print(f'Ошибка {e} товара {product}')
                                                                        if connection:
                                                                            connection.rollback()
                                                        except Exception as e:
                                                            print(e)
                            else:
                                file_path = 'data.csv'
                                cursor.execute("DELETE FROM sales_app_product WHERE shop_id = %s;", (shop['id'],))
                                # count = get_csv(shop['feeds_info'], False)

                                for feed in shop['feeds_info']:
                                    link=feed['csv_link'].replace('com/web', 'com/ru/web')+'&currency=RUB'
                                    try:
                                        try:
                                            with open(file_path, 'wb') as file:
                                                response = requests.get(link, stream=True,
                                                                        allow_redirects=True)
                                                for chunk in response.iter_content(chunk_size=100000000):
                                                    file.write(chunk)
                                        except Exception as e:
                                            print(f"Error parsing CSV: {e}")
                                            traceback.print_exc()  # Вывод полной информации о месте ошибки
                                        print('Запрос выполнен')
                                        downloaded_file_size = os.path.getsize(file_path)
                                        print(f"Скачанный размер файла: {downloaded_file_size} байт")
                                        # try:
                                        #     result = subprocess.run(['powershell', '-Command',
                                        #                              f"(Get-Content {'data.csv'} | Measure-Object -Line).Lines"],
                                        #                             stdout=subprocess.PIPE)
                                        #     line_count = int(result.stdout.decode('utf-8').strip())
                                        #
                                        #     print(f"Количество строк в файле: {line_count}")
                                        # except Exception as e:
                                        #     print(e)
                                        #     traceback.print_exc()  # Вывод полной информации о месте ошибки

                                        file_path = 'data.csv'
                                        delimiter = ';'
                                        print('Добавление товаров в таблицу')

                                        with open(file_path, 'rb') as rawdata:
                                            result = chardet.detect(
                                                rawdata.read(100000))  # Читаем часть файла для анализа
                                            encoding = result['encoding']
                                            print(f"Определенная кодировка файла: {encoding}")
                                        selected_columns = ['id', 'name', 'price', 'oldprice', 'picture', 'url',
                                                            'currencyId']
                                        for result in pd.read_csv(file_path, sep=delimiter, chunksize=10000,
                                                                  usecols=selected_columns, encoding=encoding,
                                                                  on_bad_lines='skip'):
                                            result['categoryId'] = ''
                                            correct_columns = ['id', 'name', 'categoryId', 'price', 'oldprice',
                                                               'picture', 'url', 'currencyId']
                                            result = result[correct_columns]
                                            for product in result.values:
                                                if product[7] != '':
                                                    # print(product)
                                                    # print('=================')
                                                    try:
                                                        try:
                                                            cat1 = str(product[2])
                                                        except:
                                                            cat1 = ''
                                                        try:
                                                            product[3] = product[3].replace(',', '.')
                                                        except:
                                                            pass
                                                        product[3] = re.sub(r'[^0-9.]', '', str(product[3]))
                                                        if (str(product[3]) and product[6]) not in ['NaN', 'nan', None]:
                                                            if float(product[3]) > 0:
                                                                try:
                                                                    sale = 0
                                                                    super_id = f'{str(product[0])}_%f%_{str(shop["id"])}'
                                                                    cursor.execute(
                                                                        """INSERT INTO sales_app_product (id, name, price, old_price, sale, image, url, category1, shop_id, used) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE) ON CONFLICT (id) DO NOTHING """,
                                                                        (super_id, product[1], product[3], product[4],
                                                                         sale, product[5], product[6], cat1,
                                                                         shop['id'],))
                                                                    # Фиксация транзакции
                                                                    # connection.commit()
                                                                except Exception as e:
                                                                    print(f'Ошибка {e} товара {product}')
                                                                    traceback.print_exc()  # Вывод полной информации о месте ошибки
                                                                    if connection:
                                                                        connection.rollback()
                                                    except Exception as e:
                                                        print(e)
                                                        traceback.print_exc()  # Вывод полной информации о месте ошибки
                                    except Exception as e:
                                        print(e)
                                        traceback.print_exc()

                                        # Очищаем файл после использования
                            if os.path.exists(file_path):
                                os.remove(file_path)
                            try:
                                cursor.execute(
                                    """UPDATE sales_app_shop
                                                SET last_update = %s
                                                WHERE shop_id = %s;
                                                """, (datetime.now(), shop['id']))
                                print('Обновление поиска')
                                cursor.execute("""
                                    UPDATE sales_app_product
                                    SET search = 
                                        to_tsvector('russian', COALESCE(name, '')) || 
                                        to_tsvector('russian', COALESCE(category1, ''))
                                    WHERE shop_id = %s;
                                """, (shop['id'],))
                            except Exception as e:
                                traceback.print_exc()  # Вывод полной информации о месте ошибки
                                print(f'Магазин {shop["name"]} не сохранен - {e}')
                            connection.commit()
                            if os.path.exists(file_path):
                                os.remove(file_path)
                            print(f'Товары добавлены {shop["name"]}')
                        except Exception as e:
                            traceback.print_exc()  # Вывод полной информации о месте ошибки
                            print(f'Товары не найдены {shop["name"]}  {e}')
                            not_rub_shops.append(shop['name'])
            except Exception as e:
                traceback.print_exc()  # Вывод полной информации о месте ошибки
                print(f'Магазин {shop["name"]} не найден - {e}')
                not_rub_shops.append(shop['name'])
        else:
            not_rub_shops.append(shop['name'])
    # Фиксация изменений
    # Определение данных для удаления

    connection.commit()
    cursor.close()
    connection.close()

    try:
        print('Очистка и реиндексация товаров')
        connection = psycopg2.connect(
            dbname=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD,
            host=DB_HOST,
            port=DB_PORT
        )
        connection.autocommit = True
        # Создание курсора
        cursor = connection.cursor()
        # Выполнение команды VACUUM
        # cursor.execute("VACUUM ANALYZE sales_app_shop;")
        # cursor.execute("REINDEX TABLE sales_app_shop;")
        # cursor.execute("VACUUM ANALYZE sales_app_product;")
        # cursor.execute("REINDEX TABLE sales_app_product;")
        print('Очистка и реиндексация завершены')
        # Закрытие соединения
        cursor.close()
        connection.close()
        print('Очистка и реиндексация товаров завершены')
    except Exception as e:
        print(f'{e}')

    save_store_to_xlsx(shops_names)
    return products_for_delete, not_rub_shops



def contains_russian(text):
    # Регулярное выражение для поиска русских символов
    russian_pattern = re.compile('[А-Яа-яЁё]')
    return bool(russian_pattern.search(text))

def update_database_coupons(new_data):
    # Подключение к базе данных
    conn = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = conn.cursor()
    ids_to_delete = []
    new_data = list(row for row in new_data if contains_russian(row['name']))
    try:
        # 1. Получение всех существующих записей
        cursor.execute("SELECT coupon_id, chat_id, message_id, used FROM sales_app_coupon")
        existing_data = cursor.fetchall()

        # Преобразование в словарь для удобства
        existing_data_dict = {row[0]: [row[1], row[2], row[3]] for row in existing_data}

        # 2. Сравнение и удаление устаревших записей
        new_ids = set(item['id'] for item in new_data)
        existing_ids = set(existing_data_dict.keys())

        # IDs, которые нужно удалить
        ids_to_delete = existing_ids - new_ids

        if ids_to_delete:
            cursor.execute(
                """DELETE FROM sales_app_coupon WHERE coupon_id = ANY(%s)""", (list(ids_to_delete),)
            )

        # SQL-запрос для вставки данных
        insert_query = """
                INSERT INTO sales_app_coupon (
                    coupon_id, image, shop_id, discount, name, condition, date_start, end_start, code, used, link
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (coupon_id) DO UPDATE SET link = EXCLUDED.link
                """

        for index in new_data:
            try:
                # print('Данные:', data)
                coupon_id = index['id']
                campaign_id = index['campaign']['id']
                discount = index['discount']
                name = index['name']
                condition = index['description']
                if condition is None:
                    condition = ''
                date_start = index['date_start']
                date_end = index['date_end']
                # try:
                #     ref_link, legal_info = get_campaign(index['campaign']['id'])
                # except:
                #     time.sleep(2)
                #     ref_link, legal_info = get_campaign(index['campaign']['id'])
                # image = get_banner(data['results'][index]['campaign']['id'])
                # image = convert_image(data['results'][index]['image'])
                image = ''
                used = False

                link = index['goto_link']
                try:
                    promocode = index['promocode']
                    if promocode is None:
                        promocode = 'NOT REQUIRED'
                except:
                    promocode = 'NOT REQUIRED'
            # Выполнение запроса

                cursor.execute(insert_query, (
                    coupon_id, image, campaign_id, discount, name, condition, date_start, date_end, promocode, used, link
                ))
            except:
                print(f'Не удалось сохранить {index}')


        # Сохранение изменений в базе данных
        conn.commit()
        text = f'База купонов обновлена. Удалено старых записей {(len(ids_to_delete))}'
        print(text)
    except Exception as e:
        print(f"Ошибка: {e}")
        conn.rollback()
    finally:
        conn.commit()
        cursor.close()
        conn.close()
    try:
        print('Очистка и реиндексация купонов')
        connection = psycopg2.connect(
            dbname=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD,
            host=DB_HOST,
            port=DB_PORT
        )
        connection.autocommit = True

        # Создание курсора
        cursor = connection.cursor()

        # Выполнение команды VACUUM
        cursor.execute("VACUUM ANALYZE sales_app_coupon;")
        cursor.execute("REINDEX TABLE sales_app_coupon;")
        cursor.execute("ANALYZE sales_app_coupon;")

        # Закрытие соединения
        cursor.close()
        connection.close()
        print('Очистка и реиндексация купонов завершены')
    except Exception as e:
        print(f'{e}')

    messages_for_deleting = {key: value for key, value in existing_data_dict.items() if key in ids_to_delete and value[2] is True}
    return messages_for_deleting, text




def delete_coupon_from_db(coupon_id):
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    # Выполнение запроса
    try:
        cursor.execute("""DELETE FROM sales_app_coupon WHERE id = %s; """, (coupon_id,))
        connection.commit()
    finally:
        # Закрытие курсора и подключения
        cursor.close()
        connection.close()
        print(f"Запись {coupon_id} удалена")




def select_tasks():
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    # Выполнение запроса
    try:
        cursor.execute("""SELECT chat_id, message_id, end_start, coupon_id FROM sales_app_coupon WHERE used = TRUE AND (chat_id, message_id, end_start) IS NOT NULL; """)
        tasks = cursor.fetchall()
        connection.commit()
        cursor.close()
        connection.close()
        return tasks
    except:
        return None






def upload_img_shops():
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()

    # Загружаем файл Excel
    file_path = 'Shop-Photo200.xlsx'
    wb = load_workbook(filename=file_path)
    # Выбираем активный лист (или можно указать конкретное имя листа)
    ws = wb.active

    # Перебираем строки в листе
    for row in ws.iter_rows(values_only=True):
        try:
            url_temp = 'https://thumb.cloud.mail.ru/weblink/thumb/xw1/'
            url = f"{url_temp}{row[1].split('/')[-2]}/{row[1].split('/')[-1]}"
            cursor.execute("""UPDATE sales_app_shop SET image = %s WHERE name = %s""", (url, row[0]))
        except:
            pass
    connection.commit()
    cursor.close()
    connection.close()


def categories_shop(data):
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()

    for categories in list(row['categories'] for row in data['results']):
        for category in categories:
            if category['name'] not in ['11.11', 'SME', 'New Year Festival 2024', 'eHealth', 'Black Friday RU 2023', 'Back to school WW 2023', 'Back to school']:
                try:
                    cursor.execute("""INSERT INTO sales_app_categoryshop (id, name) VALUES (%s, %s) ON CONFLICT (id) DO NOTHING""", (category['id'], category['name']))
                except:
                    pass
    connection.commit()
    cursor.close()
    connection.close()



def shop_go_out(correct_shops):
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    try:
        # Получаем старые данные (shop_id и name) из базы данных
        cursor.execute("""SELECT shop_id, name FROM sales_app_shop""")
        old_shops = cursor.fetchall()

        # Преобразуем старые магазины в словарь {shop_id: name}
        old_shops_dict = {shop[0]: shop[1] for shop in old_shops}

        # Получаем только shop_id из старых магазинов
        old_shop_ids = set(old_shops_dict.keys())

        # Преобразуем новые данные в набор с id магазинов
        correct_shop_ids = set(shop['id'] for shop in correct_shops)

        # Находим id магазинов, которых больше нет в новых данных
        go_out_shop_ids = old_shop_ids - correct_shop_ids

        # Получаем имена магазинов для этих id
        go_out_shops = [old_shops_dict[shop_id] for shop_id in go_out_shop_ids]
    except:
        go_out_shops = None
    connection.commit()
    cursor.close()
    connection.close()
    return go_out_shops


def delete_shop(name):
    connection = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )
    cursor = connection.cursor()
    try:
        cursor.execute("""SELECT shop_id FROM sales_app_shop WHERE name = %s""", (name,))
        shop_id = cursor.fetchone()[0]
        if shop_id is not None:

            cursor.execute("""DELETE FROM sales_app_coupon WHERE shop_id = %s AND used = True RETURNING chat_id, message_id;""", (shop_id,))
            coupons = cursor.fetchall()
            cursor.execute("""DELETE FROM sales_app_coupon WHERE shop_id = %s;""", (shop_id,))

            cursor.execute("""DELETE FROM sales_app_product WHERE shop_id = %s AND used = True RETURNING chat_id, message_id;""", (shop_id,))
            products = cursor.fetchall()
            cursor.execute("""DELETE FROM sales_app_product WHERE shop_id = %s;""", (shop_id,))

            cursor.execute("""DELETE FROM sales_app_shop_category WHERE shop_id = %s;""", (shop_id,))
            cursor.execute("""DELETE FROM sales_app_shop WHERE shop_id = %s;""", (shop_id,))
            connection.commit()
            cursor.close()
            connection.close()

            all_messages = coupons + products
            return all_messages
        else:
            connection.commit()
            cursor.close()
            connection.close()
            return False
    except:
        connection.commit()
        cursor.close()
        connection.close()
        return False



if __name__ == '__main__':
    upload_img_shops()